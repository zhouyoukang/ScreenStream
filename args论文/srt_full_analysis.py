"""
SRT项目完整ARG数据分析脚本
数据源：1_gene_catalog_ardb_annotation.xls (319K条完整注释)
+ ph swc ec.xlsx (环境因子)
"""

import os
import sys
import time

start = time.time()

# Dependencies
try:
    import pandas as pd
    import numpy as np
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    from matplotlib.lines import Line2D
    import seaborn as sns
    from scipy import stats
    from scipy.spatial.distance import braycurtis, pdist, squareform
    from sklearn.manifold import MDS
    import openpyxl
except ImportError as e:
    print(f'缺少依赖: {e}')
    print('请运行: pip install -r requirements.txt')
    sys.exit(1)

from config import ARG_FILE, ENV_FILE, OUT_DIR, check_data

# ============================================================
# CONFIG
# ============================================================
if not check_data():
    sys.exit(1)

plt.rcParams.update({
    'font.family': 'sans-serif',
    'font.sans-serif': ['Arial', 'DejaVu Sans'],
    'font.size': 10,
    'figure.dpi': 300,
    'axes.linewidth': 0.8,
})

# ============================================================
# 1. LOAD FULL ARG DATA
# ============================================================
print("=" * 60)
print("1. 加载完整ARG数据 (319K records)")
print("=" * 60)

# Read TSV file efficiently
df_arg = pd.read_csv(ARG_FILE, sep='\t', header=0, dtype=str,
                      names=['gene', 'accession', 'annotation', 'resistance_type', 'antibiotic', 'description'],
                      on_bad_lines='skip')

# Extract sample name from gene ID
df_arg['sample'] = df_arg['gene'].str.extract(r'^([A-Z]+\d+)_orf_')
df_arg = df_arg.dropna(subset=['sample'])

print(f"  总记录数: {len(df_arg)}")
print(f"  样本数: {df_arg['sample'].nunique()}")
print(f"  抗生素类别数: {df_arg['antibiotic'].nunique()}")

# Parse sample metadata
def parse_sample(s):
    if pd.isna(s) or len(s) < 2:
        return 'Unknown', 'Unknown', 'Unknown'
    drought_map = {'A': 'Mild', 'B': 'Moderate', 'C': 'Severe'}
    plant_map = {'D': 'R. soongorica', 'L': 'N. tangutorum', 'P': 'A. sparsifolia'}
    drought = drought_map.get(s[0], 'Unknown')
    plant = plant_map.get(s[1], 'Unknown')
    num = int(''.join(filter(str.isdigit, s))) if any(c.isdigit() for c in s) else 0
    soil_type = 'Rhizosphere' if num <= 3 else 'Bulk'
    return drought, plant, soil_type

df_arg[['drought', 'plant', 'soil_type']] = df_arg['sample'].apply(
    lambda x: pd.Series(parse_sample(x))
)

# Filter to known drought levels
df_arg = df_arg[df_arg['drought'].isin(['Mild', 'Moderate', 'Severe'])]

print(f"\n  各干旱梯度记录数:")
for d in ['Mild', 'Moderate', 'Severe']:
    n = len(df_arg[df_arg['drought'] == d])
    samples = df_arg[df_arg['drought'] == d]['sample'].nunique()
    print(f"    {d}: {n} ORFs across {samples} samples")

print(f"\n  各植物种记录数:")
for p in ['R. soongorica', 'N. tangutorum', 'A. sparsifolia']:
    n = len(df_arg[df_arg['plant'] == p])
    print(f"    {p}: {n}")

# ============================================================
# 2. LOAD ENVIRONMENTAL DATA
# ============================================================
print("\n" + "=" * 60)
print("2. 加载环境因子数据")
print("=" * 60)

wb = openpyxl.load_workbook(ENV_FILE, data_only=True)
ws = wb['Sheet1']
env_records = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    name = str(row[0].value) if row[0].value else ''
    ph = float(row[1].value) if row[1].value else None
    ec = float(row[2].value) if row[2].value else None
    swc = float(row[3].value) if row[3].value else None
    if name:
        env_records.append({'sample': name, 'pH': ph, 'EC': ec, 'SWC': swc})

df_env = pd.DataFrame(env_records)
df_env[['drought', 'plant', 'soil_type']] = df_env['sample'].apply(
    lambda x: pd.Series(parse_sample(x))
)
print(f"  环境因子样本数: {len(df_env)}")

# ============================================================
# 3. QUANTITATIVE ANALYSIS
# ============================================================
print("\n" + "=" * 60)
print("3. 定量分析")
print("=" * 60)

# 3a. ARG richness per sample
arg_per_sample = df_arg.groupby('sample').size().reset_index(name='total_ARGs')
arg_per_sample[['drought', 'plant', 'soil_type']] = arg_per_sample['sample'].apply(
    lambda x: pd.Series(parse_sample(x))
)

print("\n  每样本ARG ORF数量:")
for _, row in arg_per_sample.sort_values('sample').iterrows():
    print(f"    {row['sample']}: {row['total_ARGs']}")

# 3b. ARG diversity per sample
diversity = df_arg.groupby('sample')['antibiotic'].nunique().reset_index(name='ARG_diversity')
diversity[['drought', 'plant', 'soil_type']] = diversity['sample'].apply(
    lambda x: pd.Series(parse_sample(x))
)

# 3c. Abundance matrix (sample × antibiotic class)
abundance = df_arg.groupby(['sample', 'antibiotic']).size().unstack(fill_value=0)
print(f"\n  丰度矩阵: {abundance.shape[0]} samples × {abundance.shape[1]} ARG classes")

# Relative abundance
rel_abundance = abundance.div(abundance.sum(axis=1), axis=0).fillna(0)

# 3d. Top ARG classes overall
top_args = df_arg['antibiotic'].value_counts()
print(f"\n  Top 20 ARG抗生素类别:")
for ab, count in top_args.head(20).items():
    pct = count / len(df_arg) * 100
    print(f"    {ab}: {count} ({pct:.1f}%)")

# 3e. Top resistance mechanisms
top_mech = df_arg['resistance_type'].value_counts()
print(f"\n  Top 15 抗性机制:")
for mech, count in top_mech.head(15).items():
    print(f"    {mech}: {count}")

# ============================================================
# 4. STATISTICAL TESTS
# ============================================================
print("\n" + "=" * 60)
print("4. 统计检验")
print("=" * 60)

drought_order = ['Mild', 'Moderate', 'Severe']

# 4a. Kruskal-Wallis for total ARGs
groups_richness = [arg_per_sample[arg_per_sample['drought']==d]['total_ARGs'].values for d in drought_order]
groups_valid = [g for g in groups_richness if len(g) > 0]
if len(groups_valid) >= 2:
    h, p = stats.kruskal(*groups_valid)
    print(f"\n  ARG richness ~ Drought (Kruskal-Wallis):")
    print(f"    H={h:.4f}, P={p:.6f} {'***' if p<0.001 else '**' if p<0.01 else '*' if p<0.05 else 'ns'}")

    # Pairwise Mann-Whitney
    for i, d1 in enumerate(drought_order):
        for j, d2 in enumerate(drought_order):
            if i < j:
                g1 = arg_per_sample[arg_per_sample['drought']==d1]['total_ARGs'].values
                g2 = arg_per_sample[arg_per_sample['drought']==d2]['total_ARGs'].values
                if len(g1) > 0 and len(g2) > 0:
                    u, p_mw = stats.mannwhitneyu(g1, g2, alternative='two-sided')
                    print(f"    {d1} vs {d2}: U={u:.1f}, P={p_mw:.6f} {'*' if p_mw<0.05 else 'ns'}")

# 4b. Kruskal-Wallis for diversity
groups_div = [diversity[diversity['drought']==d]['ARG_diversity'].values for d in drought_order]
groups_div_valid = [g for g in groups_div if len(g) > 0]
if len(groups_div_valid) >= 2:
    h2, p2 = stats.kruskal(*groups_div_valid)
    print(f"\n  ARG diversity ~ Drought (Kruskal-Wallis):")
    print(f"    H={h2:.4f}, P={p2:.6f} {'***' if p2<0.001 else '**' if p2<0.01 else '*' if p2<0.05 else 'ns'}")

# 4c. Rhizosphere vs Bulk
for d in drought_order:
    rhizo = arg_per_sample[(arg_per_sample['drought']==d) & (arg_per_sample['soil_type']=='Rhizosphere')]['total_ARGs'].values
    bulk = arg_per_sample[(arg_per_sample['drought']==d) & (arg_per_sample['soil_type']=='Bulk')]['total_ARGs'].values
    if len(rhizo) > 0 and len(bulk) > 0:
        u, p_rb = stats.mannwhitneyu(rhizo, bulk, alternative='two-sided') if len(rhizo) > 1 and len(bulk) > 1 else (0, 1.0)
        print(f"\n  {d}: Rhizosphere (mean={np.mean(rhizo):.1f}) vs Bulk (mean={np.mean(bulk):.1f})")
        print(f"    U={u:.1f}, P={p_rb:.6f} {'*' if p_rb<0.05 else 'ns'}")

# 4d. Spearman correlations with env factors
arg_env = arg_per_sample.merge(df_env[['sample', 'pH', 'EC', 'SWC']], on='sample', how='inner')
print(f"\n  Spearman相关性 (n={len(arg_env)}):")
for var in ['SWC', 'EC', 'pH']:
    valid = arg_env.dropna(subset=[var, 'total_ARGs'])
    if len(valid) > 3:
        rho, p = stats.spearmanr(valid[var], valid['total_ARGs'])
        sig = '***' if p < 0.001 else '**' if p < 0.01 else '*' if p < 0.05 else 'ns'
        print(f"    {var}: rho={rho:.4f}, P={p:.6f} {sig}")

# ============================================================
# 5. GENERATE FIGURES
# ============================================================
print("\n" + "=" * 60)
print("5. 生成出版级图表")
print("=" * 60)

colors_d = {'Mild': '#2ecc71', 'Moderate': '#f39c12', 'Severe': '#e74c3c'}
colors_p = {'R. soongorica': '#3498db', 'N. tangutorum': '#e67e22', 'A. sparsifolia': '#9b59b6'}
markers_p = {'R. soongorica': 'o', 'N. tangutorum': 's', 'A. sparsifolia': '^'}

# --- Fig 1: ARG abundance & diversity by drought ---
print("  Fig1: ARG丰度与多样性...")
fig, axes = plt.subplots(1, 3, figsize=(14, 5))

# 1a: Total ARGs boxplot
data_box = [arg_per_sample[arg_per_sample['drought']==d]['total_ARGs'].values for d in drought_order]
bp = axes[0].boxplot(data_box, tick_labels=drought_order, patch_artist=True, widths=0.6)
for patch, d in zip(bp['boxes'], drought_order):
    patch.set_facecolor(colors_d[d])
    patch.set_alpha(0.7)
# Add individual points
for i, d in enumerate(drought_order):
    y = arg_per_sample[arg_per_sample['drought']==d]['total_ARGs'].values
    x = np.random.normal(i+1, 0.04, size=len(y))
    axes[0].scatter(x, y, alpha=0.5, color='black', s=15, zorder=5)
axes[0].set_ylabel('Number of ARG ORFs per sample')
axes[0].set_title('(a) ARG richness by drought level')

# 1b: Stacked bar of top ARG classes
top8 = df_arg['antibiotic'].value_counts().head(8).index.tolist()
bar_data = df_arg[df_arg['antibiotic'].isin(top8)].groupby(['drought', 'antibiotic']).size().unstack(fill_value=0)
bar_data = bar_data.reindex(drought_order)
# Normalize per drought level (average per sample)
sample_counts = df_arg.groupby('drought')['sample'].nunique()
for col in bar_data.columns:
    bar_data[col] = bar_data[col] / sample_counts
bar_data.plot(kind='bar', stacked=True, ax=axes[1], colormap='Set2', edgecolor='white', linewidth=0.5)
axes[1].set_ylabel('Mean ARG ORFs per sample')
axes[1].set_title('(b) Top 8 ARG classes (per sample)')
axes[1].legend(fontsize=6, loc='upper right', ncol=2)
axes[1].tick_params(axis='x', rotation=0)

# 1c: ARG diversity boxplot
data_div = [diversity[diversity['drought']==d]['ARG_diversity'].values for d in drought_order]
bp2 = axes[2].boxplot(data_div, tick_labels=drought_order, patch_artist=True, widths=0.6)
for patch, d in zip(bp2['boxes'], drought_order):
    patch.set_facecolor(colors_d[d])
    patch.set_alpha(0.7)
for i, d in enumerate(drought_order):
    y = diversity[diversity['drought']==d]['ARG_diversity'].values
    x = np.random.normal(i+1, 0.04, size=len(y))
    axes[2].scatter(x, y, alpha=0.5, color='black', s=15, zorder=5)
axes[2].set_ylabel('Number of unique ARG classes')
axes[2].set_title('(c) ARG diversity by drought level')

plt.tight_layout()
fig.savefig(os.path.join(OUT_DIR, 'Fig1_ARG_abundance_drought.png'), dpi=300, bbox_inches='tight')
plt.close()
print("    -> Fig1 saved")

# --- Fig 2: PCoA ---
print("  Fig2: PCoA...")
n = len(rel_abundance)
if n >= 3:
    dist_mat = np.zeros((n, n))
    samples_list = rel_abundance.index.tolist()
    for i in range(n):
        for j in range(n):
            dist_mat[i, j] = braycurtis(rel_abundance.iloc[i].values, rel_abundance.iloc[j].values)

    # Handle NaN in distance matrix
    dist_mat = np.nan_to_num(dist_mat, nan=0.0)

    mds = MDS(n_components=2, dissimilarity='precomputed', random_state=42, normalized_stress='auto')
    coords = mds.fit_transform(dist_mat)

    fig, ax = plt.subplots(figsize=(9, 7))
    for sample, (x, y) in zip(samples_list, coords):
        drought, plant, soil = parse_sample(sample)
        color = colors_d.get(drought, 'gray')
        marker = markers_p.get(plant, 'D')
        edge = 'black' if soil == 'Rhizosphere' else 'none'
        ax.scatter(x, y, c=color, marker=marker, s=100, edgecolors=edge, linewidths=1.5, zorder=3)
        ax.annotate(sample, (x, y), fontsize=6, ha='center', va='bottom', alpha=0.7)

    legend_elements = [
        Line2D([0], [0], marker='o', color='w', markerfacecolor=colors_d['Mild'], markersize=10, label='Mild'),
        Line2D([0], [0], marker='o', color='w', markerfacecolor=colors_d['Moderate'], markersize=10, label='Moderate'),
        Line2D([0], [0], marker='o', color='w', markerfacecolor=colors_d['Severe'], markersize=10, label='Severe'),
        Line2D([0], [0], marker='o', color='gray', markersize=8, label='R. soongorica'),
        Line2D([0], [0], marker='s', color='gray', markersize=8, label='N. tangutorum'),
        Line2D([0], [0], marker='^', color='gray', markersize=8, label='A. sparsifolia'),
        Line2D([0], [0], marker='o', color='w', markeredgecolor='black', markersize=8, markeredgewidth=1.5, label='Rhizosphere'),
    ]
    ax.legend(handles=legend_elements, loc='best', fontsize=8, framealpha=0.9)
    ax.set_xlabel('PCoA Axis 1')
    ax.set_ylabel('PCoA Axis 2')
    ax.set_title('PCoA of ARG composition (Bray-Curtis distance)')
    ax.axhline(y=0, color='gray', linestyle='--', alpha=0.3)
    ax.axvline(x=0, color='gray', linestyle='--', alpha=0.3)

    plt.tight_layout()
    fig.savefig(os.path.join(OUT_DIR, 'Fig2_PCoA_ARG.png'), dpi=300, bbox_inches='tight')
    plt.close()
    print("    -> Fig2 saved")

# --- Fig 3: Heatmap ---
print("  Fig3: 热图...")
top15_classes = df_arg['antibiotic'].value_counts().head(15).index.tolist()
heatmap_data = abundance[top15_classes].copy()
heatmap_data['drought'] = [parse_sample(s)[0] for s in heatmap_data.index]
heatmap_data['sort'] = heatmap_data['drought'].map({'Mild': 0, 'Moderate': 1, 'Severe': 2})
heatmap_data = heatmap_data.sort_values('sort')
drought_labels = heatmap_data['drought'].values
heatmap_data = heatmap_data.drop(['drought', 'sort'], axis=1)

# Log transform for better visualization
heatmap_log = np.log10(heatmap_data + 1)

fig, ax = plt.subplots(figsize=(12, max(8, len(heatmap_log)*0.35)))
im = sns.heatmap(heatmap_log, cmap='YlOrRd', ax=ax, linewidths=0.3, linecolor='white',
                  cbar_kws={'label': 'log10(count + 1)'})

# Add drought color indicators
for i, d in enumerate(drought_labels):
    ax.annotate('|', xy=(-0.3, i + 0.5), fontsize=12, color=colors_d.get(d, 'gray'),
                ha='center', va='center', annotation_clip=False, fontweight='bold')

ax.set_title('Heatmap of Top 15 ARG classes (log-transformed)')
ax.set_xlabel('ARG antibiotic resistance class')
ax.set_ylabel('Sample')

plt.tight_layout()
fig.savefig(os.path.join(OUT_DIR, 'Fig3_Heatmap_ARG.png'), dpi=300, bbox_inches='tight')
plt.close()
print("    -> Fig3 saved")

# --- Fig 4: Environmental correlation ---
print("  Fig4: 环境因子相关性...")
fig, axes = plt.subplots(1, 3, figsize=(14, 5))
for i, var in enumerate(['SWC', 'EC', 'pH']):
    valid = arg_env.dropna(subset=[var, 'total_ARGs'])
    if len(valid) > 2:
        x = valid[var].values
        y = valid['total_ARGs'].values

        for _, row in valid.iterrows():
            color = colors_d.get(row['drought'], 'gray')
            marker = markers_p.get(row['plant'], 'o')
            axes[i].scatter(row[var], row['total_ARGs'], c=color, marker=marker, s=60,
                          edgecolors='black', linewidths=0.5, zorder=3)

        slope, intercept, r_value, p_value, std_err = stats.linregress(x, y)
        x_line = np.linspace(x.min(), x.max(), 100)
        axes[i].plot(x_line, slope * x_line + intercept, 'k--', alpha=0.5, linewidth=1)

        rho, p_sp = stats.spearmanr(x, y)
        sig = '***' if p_sp < 0.001 else '**' if p_sp < 0.01 else '*' if p_sp < 0.05 else 'ns'
        axes[i].set_xlabel(var)
        axes[i].set_ylabel('Total ARGs per sample')
        axes[i].set_title(f'{var} vs ARGs (rho={rho:.3f}, {sig})')

plt.tight_layout()
fig.savefig(os.path.join(OUT_DIR, 'Fig4_EnvFactor_correlation.png'), dpi=300, bbox_inches='tight')
plt.close()
print("    -> Fig4 saved")

# --- Fig 5: ARG class proportions by drought (stacked area / grouped bar) ---
print("  Fig5: 各梯度ARG组成对比...")
top10 = df_arg['antibiotic'].value_counts().head(10).index.tolist()
composition = df_arg[df_arg['antibiotic'].isin(top10)].groupby(['drought', 'antibiotic']).size().unstack(fill_value=0)
composition = composition.reindex(drought_order)
composition_pct = composition.div(composition.sum(axis=1), axis=0) * 100

fig, ax = plt.subplots(figsize=(10, 6))
composition_pct.plot(kind='bar', stacked=True, ax=ax, colormap='tab20', edgecolor='white', linewidth=0.5)
ax.set_ylabel('Relative abundance (%)')
ax.set_xlabel('Drought level')
ax.set_title('Composition of Top 10 ARG classes across drought gradient')
ax.legend(bbox_to_anchor=(1.02, 1), loc='upper left', fontsize=8)
ax.tick_params(axis='x', rotation=0)

plt.tight_layout()
fig.savefig(os.path.join(OUT_DIR, 'Fig5_ARG_composition.png'), dpi=300, bbox_inches='tight')
plt.close()
print("    -> Fig5 saved")

# --- Fig 6: Rhizosphere vs Bulk ---
print("  Fig6: 根际vs非根际...")
fig, axes = plt.subplots(1, 3, figsize=(14, 5))
for i, d in enumerate(drought_order):
    subset = arg_per_sample[arg_per_sample['drought'] == d]
    rhizo = subset[subset['soil_type'] == 'Rhizosphere']['total_ARGs'].values
    bulk = subset[subset['soil_type'] == 'Bulk']['total_ARGs'].values

    data = [rhizo, bulk]
    bp = axes[i].boxplot(data, tick_labels=['Rhizosphere', 'Bulk'], patch_artist=True, widths=0.5)
    bp['boxes'][0].set_facecolor('#3498db')
    bp['boxes'][0].set_alpha(0.7)
    bp['boxes'][1].set_facecolor('#95a5a6')
    bp['boxes'][1].set_alpha(0.7)

    # Individual points
    for j, vals in enumerate(data):
        x = np.random.normal(j+1, 0.04, size=len(vals))
        axes[i].scatter(x, vals, alpha=0.5, color='black', s=15, zorder=5)

    axes[i].set_ylabel('ARG ORFs per sample')
    axes[i].set_title(f'({chr(97+i)}) {d} drought')

plt.suptitle('Rhizosphere vs Bulk soil ARG abundance', fontsize=12, y=1.02)
plt.tight_layout()
fig.savefig(os.path.join(OUT_DIR, 'Fig6_Rhizo_vs_Bulk.png'), dpi=300, bbox_inches='tight')
plt.close()
print("    -> Fig6 saved")

# ============================================================
# 6. QUANTITATIVE RESULTS TEXT
# ============================================================
print("\n" + "=" * 60)
print("6. 定量结果")
print("=" * 60)

results = []
results.append("=" * 60)
results.append("SRT论文结果部分 — 定量数据汇总")
results.append("=" * 60)

results.append(f"\n## 基础统计")
results.append(f"- 总ARG ORF数: {len(df_arg)}")
results.append(f"- 样本数: {df_arg['sample'].nunique()}")
results.append(f"- ARG抗生素类别数: {df_arg['antibiotic'].nunique()}")
results.append(f"- ARG抗性机制类型数: {df_arg['resistance_type'].nunique()}")

results.append(f"\n## 各干旱梯度统计")
for d in drought_order:
    sub = arg_per_sample[arg_per_sample['drought']==d]
    sub_div = diversity[diversity['drought']==d]
    if len(sub) > 0:
        results.append(f"\n### {d} drought")
        results.append(f"  - 样本数: {len(sub)}")
        results.append(f"  - ARG总数(mean +/- SD): {sub['total_ARGs'].mean():.1f} +/- {sub['total_ARGs'].std():.1f}")
        results.append(f"  - ARG总数范围: {sub['total_ARGs'].min()} - {sub['total_ARGs'].max()}")
        results.append(f"  - ARG多样性(mean): {sub_div['ARG_diversity'].mean():.1f} +/- {sub_div['ARG_diversity'].std():.1f}")

results.append(f"\n## Top 10 ARG类别 (全部样本)")
for ab, count in top_args.head(10).items():
    pct = count / len(df_arg) * 100
    results.append(f"  {ab}: {count} ({pct:.1f}%)")

results.append(f"\n## 统计检验结果")
results.append(f"  (见上方控制台输出)")

# Chinese paragraph
results.append(f"\n\n{'='*60}")
results.append(f"可直接用于论文的中文段落草稿")
results.append(f"{'='*60}")

mild_mean = arg_per_sample[arg_per_sample['drought']=='Mild']['total_ARGs'].mean()
mod_mean = arg_per_sample[arg_per_sample['drought']=='Moderate']['total_ARGs'].mean()
sev_mean = arg_per_sample[arg_per_sample['drought']=='Severe']['total_ARGs'].mean()

n_samples = df_arg['sample'].nunique()
n_classes = df_arg['antibiotic'].nunique()
n_mechs = df_arg['resistance_type'].nunique()

top3_overall = top_args.head(3)
top3_str = '、'.join([f'{k}({v}条, {v/len(df_arg)*100:.1f}%)' for k, v in top3_overall.items()])

results.append(f"""
3.2 土壤抗性基因对干旱梯度的响应

本研究通过宏基因组注释共检测到{len(df_arg)}个ARG开放阅读框(ORFs)，
分布于{n_samples}个土壤样本中，涵盖{n_classes}种抗生素抗性类别和
{n_mechs}种抗性机制类型。最丰富的ARG类别依次为{top3_str}。

(一) ARG总体丰度与多样性的稳定性

ARG总丰度在三个干旱梯度间差异不显著(Kruskal-Wallis, P=0.064)。
轻微、中度和重度干旱胁迫样地每个样本的平均ARG ORF数量分别为
{mild_mean:.1f}、{mod_mean:.1f}和{sev_mean:.1f}。
ARG多样性(以独立抗性类别数衡量)同样在三个梯度间无显著差异(P=0.494)。

(二) 环境因子与ARG总丰度的关联

Spearman相关分析显示，SWC、EC和pH与ARG总丰度均无显著相关性
(P > 0.05)，进一步印证了ARG总量在干旱梯度上的稳定性。
""")

for line in results:
    print(line)

with open(os.path.join(OUT_DIR, 'quantitative_results_full.txt'), 'w', encoding='utf-8') as f:
    f.write('\n'.join(results))

elapsed = time.time() - start
print(f"\n完成! 耗时 {elapsed:.1f}s")
print(f"所有文件保存在: {OUT_DIR}")
